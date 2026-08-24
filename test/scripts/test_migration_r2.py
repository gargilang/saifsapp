from datetime import date, datetime
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from scripts.migration_r2 import (
    ALLOWED_JENIS,
    DateNormalizationError,
    NormalizedTimeline,
    build_dataset,
    date_candidates,
    normalize_timeline,
    validate_dataset,
    verify_preview,
    write_preview,
)


ROOT = Path(__file__).resolve().parents[2]
R2 = ROOT / "ref" / "DAFTAR KREDIT BARANG r2.xlsx"
LEGACY = ROOT / "ref" / "PREVIEW_MIGRASI.xlsx"


class DateNormalizationTest(unittest.TestCase):
    def test_ambiguous_excel_date_has_stored_and_swapped_candidates(self):
        self.assertEqual(
            date_candidates(datetime(2021, 8, 1)),
            (date(2021, 8, 1), date(2021, 1, 8)),
        )

    def test_order_is_swapped_when_installments_would_precede_it(self):
        result = normalize_timeline(
            datetime(2021, 8, 1),
            [datetime(2021, 2, 24), datetime(2021, 3, 29)],
            cutoff=date(2026, 8, 24),
        )
        self.assertEqual(result.order, date(2021, 1, 8))
        self.assertEqual(result.payments, (date(2021, 2, 24), date(2021, 3, 29)))

    def test_late_ambiguous_installment_prefers_fewer_sequence_inversions(self):
        result = normalize_timeline(
            datetime(2020, 12, 30),
            [datetime(2021, 5, 31), datetime(2021, 1, 7)],
            cutoff=date(2026, 8, 24),
        )
        self.assertEqual(result.payments[-1], date(2021, 7, 1))

    def test_future_us_dates_offer_9_and_11_june_candidates(self):
        self.assertIn(date(2026, 6, 9), date_candidates(datetime(2026, 9, 6)))
        self.assertIn(date(2026, 6, 11), date_candidates(datetime(2026, 11, 6)))

    def test_payment_still_before_order_is_reported(self):
        result = normalize_timeline(
            datetime(2026, 8, 20),
            [datetime(2025, 12, 31)],
            cutoff=date(2026, 8, 24),
        )
        self.assertIn("payment_before_order", result.issues)

    def test_missing_date_raises_instead_of_using_january_fallback(self):
        with self.assertRaises(DateNormalizationError):
            normalize_timeline(None, [], cutoff=date(2026, 8, 24))

    def test_two_digit_indonesian_year_is_supported(self):
        self.assertEqual(
            date_candidates("30/1/21"),
            (date(2021, 1, 30),),
        )

    def test_successful_correction_is_not_counted_as_residual_warning(self):
        timeline = NormalizedTimeline(
            order=date(2021, 1, 8),
            payments=(date(2021, 2, 24),),
            corrections=(
                {
                    "field": "order",
                    "index": 0,
                    "raw": date(2021, 8, 1),
                    "normalized": date(2021, 1, 8),
                    "reason": "day_month_swap",
                },
            ),
        )
        report = validate_dataset([timeline], cutoff=date(2026, 8, 24))
        self.assertEqual(report.transactions_with_date_warnings, 0)

    def test_future_order_is_a_hard_validation_error(self):
        timeline = NormalizedTimeline(
            order=date(2026, 9, 1),
            payments=(),
            issues=("future_date",),
        )
        report = validate_dataset([timeline], cutoff=date(2026, 8, 24))
        self.assertFalse(report.ok)
        self.assertIn("future_order", report.errors[0])

    def test_sequence_inversion_alone_is_not_a_residual_date_warning(self):
        timeline = NormalizedTimeline(
            order=date(2021, 1, 1),
            payments=(date(2021, 3, 1), date(2021, 2, 1)),
            issues=("payment_sequence_inversion",),
        )
        report = validate_dataset([timeline], cutoff=date(2026, 8, 24))
        self.assertEqual(report.transactions_with_date_warnings, 0)


class RealWorkbookDatasetTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.dataset = build_dataset(R2, LEGACY, cutoff=date(2026, 8, 24))

    def test_real_r2_control_totals(self):
        dataset = self.dataset
        self.assertEqual(len(dataset.customers), 46)
        self.assertEqual(len(dataset.transactions), 249)
        self.assertEqual(len(dataset.payments), 1069)
        self.assertEqual(sum(row.harga_beli for row in dataset.transactions), 877_769_000)
        self.assertEqual(sum(row.harga_jual for row in dataset.transactions), 971_417_500)
        self.assertEqual(sum(row.jumlah for row in dataset.payments), 854_692_500)
        self.assertEqual(sum(row.sisa for row in dataset.transactions), 116_725_000)
        self.assertEqual(dataset.validation.future_payment_count, 0)
        self.assertEqual(dataset.validation.payment_before_order_count, 19)
        self.assertEqual(dataset.validation.transactions_with_date_warnings, 17)

    def test_explicit_item_fixes_and_allowed_jenis(self):
        expected = {
            10: "AC Mobil",
            46: "Pinjaman",
            78: "Freezer",
            89: "Spare Part Mobil",
            145: "Lain-lain",
            146: "Lain-lain",
            245: "Shockbreaker",
        }
        by_number = {row.no: row for row in self.dataset.transactions}
        self.assertEqual(
            {number: by_number[number].item for number in expected},
            expected,
        )
        self.assertTrue(
            all(row.jenis in ALLOWED_JENIS for row in self.dataset.transactions)
        )

    def test_preview_has_five_reviewable_sheets_and_reloads_validly(self):
        with TemporaryDirectory() as directory:
            output = Path(directory) / "preview.xlsx"
            write_preview(self.dataset, output)
            workbook = load_workbook(output, data_only=True)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "TRANSAKSI",
                    "CUSTOMER",
                    "CICILAN",
                    "SALDO_SUMBER_DANA",
                    "VALIDASI",
                ],
            )
            funds = list(
                workbook["SALDO_SUMBER_DANA"].iter_rows(
                    min_row=3, max_row=4, values_only=True
                )
            )
            self.assertEqual(funds, [("Sandi", 36_100_000), ("Ika", 80_625_000)])
            report = verify_preview(output)
            self.assertTrue(report.ok)
            self.assertEqual(report.payment_before_order_count, 19)
            self.assertEqual(report.transactions_with_date_warnings, 17)


if __name__ == "__main__":
    unittest.main()
