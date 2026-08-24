"""Pure helpers for normalizing dates in the r2 migration workbook.

The workbook was entered with an Indonesian ``dd/mm/yyyy`` intent, while
Excel may have stored an ambiguous value using the US interpretation.  Date
normalization therefore keeps both interpretations where possible and picks
the deterministic path with the fewest chronology problems.
"""

from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import date, datetime, time
from itertools import product
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils.datetime import from_excel
except ImportError:  # pragma: no cover - openpyxl is supplied by the migration env
    Workbook = None
    load_workbook = None
    from_excel = None


ITEM_FIXES = {
    10: "AC Mobil",
    46: "Pinjaman",
    78: "Freezer",
    89: "Spare Part Mobil",
    145: "Lain-lain",
    146: "Lain-lain",
    245: "Shockbreaker",
}
ALLOWED_JENIS = {"barang", "pinjaman", "investasi", "jasa/servis", "modal usaha"}
OPENING_FUNDS = {"Sandi": 36_100_000, "Ika": 80_625_000}
EXPECTED_CONTROLS = {
    "customers": 46,
    "transactions": 249,
    "payments": 1069,
    "harga_beli": 877_769_000,
    "harga_jual": 971_417_500,
    "payments_total": 854_692_500,
    "sisa": 116_725_000,
}


def _format_rupiah(value: int) -> str:
    return f"Rp{value:,}".replace(",", ".")


class DateNormalizationError(ValueError):
    """Raised when a required workbook date cannot be interpreted."""


def _coerce_string(value: str) -> date:
    value = value.strip()
    if not value:
        raise DateNormalizationError("empty date value")

    # An explicit slash-separated value is the client's intended Indonesian
    # format.  ISO values are accepted for generated previews and cutoffs.
    for fmt in (
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%Y/%m/%d",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(value).date()
    except ValueError as exc:
        raise DateNormalizationError(f"invalid date value: {value!r}") from exc


def coerce_excel_date(value: object) -> date:
    """Convert an openpyxl/Excel date value to a ``date``.

    No placeholder date is produced for null or malformed values.  Numeric
    values are interpreted as Excel serials using openpyxl's epoch handling.
    """

    if value is None or isinstance(value, bool):
        raise DateNormalizationError("missing date value")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return _coerce_string(value)
    if isinstance(value, (int, float)):
        if from_excel is None:
            raise DateNormalizationError("openpyxl is required for Excel serial dates")
        try:
            converted = from_excel(value)
        except (TypeError, ValueError, OverflowError) as exc:
            raise DateNormalizationError(f"invalid Excel serial date: {value!r}") from exc
        if isinstance(converted, datetime):
            return converted.date()
        if isinstance(converted, date):
            return converted
        if isinstance(converted, time):
            raise DateNormalizationError(f"Excel serial has no date: {value!r}")
    raise DateNormalizationError(f"unsupported date value: {value!r}")


def date_candidates(value: object) -> tuple[date, ...]:
    """Return stored and, when valid, day/month-swapped date candidates."""

    stored = coerce_excel_date(value)
    candidates = [stored]
    if stored.day <= 12 and stored.month <= 12 and stored.day != stored.month:
        swapped = date(stored.year, stored.day, stored.month)
        candidates.append(swapped)
    return tuple(dict.fromkeys(candidates))


def _pairwise(values: Sequence[date]) -> Iterable[tuple[date, date]]:
    return zip(values, values[1:])


def _score(path: tuple[date, ...], raw: tuple[date, ...], cutoff: date) -> tuple[int, int, int, int, tuple[date, ...]]:
    purchase_date, *payment_dates = path
    future = sum(value > cutoff for value in path)
    before_order = sum(value < purchase_date for value in payment_dates)
    inversions = sum(a > b for a, b in _pairwise(payment_dates))
    swaps = sum(chosen != original for chosen, original in zip(path, raw))
    return future, before_order, inversions, swaps, path


def _issue_names(metrics: tuple[int, int, int, int]) -> tuple[str, ...]:
    future, before_order, inversions, _swaps = metrics
    issues: list[str] = []
    if future:
        issues.append("future_date")
    if before_order:
        issues.append("payment_before_order")
    if inversions:
        issues.append("payment_sequence_inversion")
    return tuple(issues)


def _correction_entries(raw: tuple[date, ...], chosen: tuple[date, ...]) -> tuple[dict[str, Any], ...]:
    corrections: list[dict[str, Any]] = []
    for index, (original, normalized) in enumerate(zip(raw, chosen)):
        if original == normalized:
            continue
        field = "order" if index == 0 else "payment"
        corrections.append(
            {
                "field": field,
                "index": index,
                "raw": original,
                "normalized": normalized,
                "reason": "day_month_swap",
            }
        )
    return tuple(corrections)


@dataclass(frozen=True)
class TimelineCandidate:
    """One candidate path retained for the validation sheet."""

    path: tuple[date, ...]
    score: tuple[int, int, int, int]
    reasons: tuple[str, ...]


@dataclass(frozen=True)
class NormalizedTimeline:
    """Selected timeline plus all evidence used to select it."""

    order: date
    payments: tuple[date, ...]
    issues: tuple[str, ...] = ()
    corrections: tuple[dict[str, Any], ...] = ()
    raw: tuple[date, ...] = ()
    candidates: tuple[tuple[date, ...], ...] = ()
    scores: tuple[tuple[int, int, int, int], ...] = ()
    candidate_details: tuple[TimelineCandidate, ...] = ()
    selected_score: tuple[int, int, int, int] = ()

    @property
    def candidate_paths(self) -> tuple[tuple[date, ...], ...]:
        return self.candidates

    @property
    def candidate_scores(self) -> tuple[tuple[int, int, int, int], ...]:
        return self.scores

    @property
    def correction_reasons(self) -> tuple[str, ...]:
        return tuple(entry["reason"] for entry in self.corrections)


def normalize_timeline(order: object, payments: Iterable[object], cutoff: object) -> NormalizedTimeline:
    """Choose a date interpretation using chronology-first deterministic scoring."""

    try:
        payment_values = tuple(payments)
    except TypeError as exc:
        raise DateNormalizationError("payments must be iterable") from exc

    raw = (coerce_excel_date(order), *(coerce_excel_date(payment) for payment in payment_values))
    cutoff_date = coerce_excel_date(cutoff)
    candidate_sets = (date_candidates(order), *(date_candidates(payment) for payment in payment_values))
    paths = tuple(product(*candidate_sets))

    scored = tuple((_score(path, raw, cutoff_date), path) for path in paths)
    chosen_score_with_path, chosen = min(scored, key=lambda item: item[0])
    selected_metrics = chosen_score_with_path[:4]
    details = tuple(
        TimelineCandidate(path=path, score=score[:4], reasons=_issue_names(score[:4]))
        for score, path in scored
    )
    return NormalizedTimeline(
        order=chosen[0],
        payments=tuple(chosen[1:]),
        issues=_issue_names(selected_metrics),
        corrections=_correction_entries(raw, chosen),
        raw=raw,
        candidates=tuple(path for path in paths),
        scores=tuple(score[:4] for score, _path in scored),
        candidate_details=details,
        selected_score=selected_metrics,
    )


@dataclass(frozen=True)
class ValidationReport:
    """Date-focused validation output used by preview generation."""

    ok: bool
    errors: tuple[str, ...] = ()
    warnings: tuple[str, ...] = ()
    issues: tuple[str, ...] = ()
    future_payment_count: int = 0
    payment_before_order_count: int = 0
    transactions_with_date_warnings: int = 0
    summary: str = ""


def _dataset_timelines(dataset: object) -> tuple[object, ...]:
    if isinstance(dataset, Mapping):
        for key in ("timelines", "transactions"):
            if key in dataset:
                return tuple(dataset[key])
        return ()
    for key in ("timelines", "transactions"):
        value = getattr(dataset, key, None)
        if value is not None:
            return tuple(value)
    if isinstance(dataset, Iterable) and not isinstance(dataset, (str, bytes)):
        return tuple(dataset)
    return ()


def _timeline_from_row(row: object) -> NormalizedTimeline | None:
    if isinstance(row, NormalizedTimeline):
        return row
    if isinstance(row, Mapping):
        candidate = row.get("timeline")
        if isinstance(candidate, NormalizedTimeline):
            return candidate
        order = row.get("order", row.get("tanggal"))
        payments = row.get("payments", ())
    else:
        candidate = getattr(row, "timeline", None)
        if isinstance(candidate, NormalizedTimeline):
            return candidate
        order = getattr(row, "order", getattr(row, "tanggal", None))
        payments = getattr(row, "payments", ())
    if isinstance(order, date) and all(isinstance(payment, date) for payment in payments):
        payment_dates = tuple(payments)
        before = sum(payment < order for payment in payment_dates)
        inversions = sum(a > b for a, b in _pairwise(payment_dates))
        issues = tuple(
            name
            for name, count in (
                ("payment_before_order", before),
                ("payment_sequence_inversion", inversions),
            )
            if count
        )
        return NormalizedTimeline(order=order, payments=payment_dates, issues=issues)
    return None


def validate_dataset(dataset: object, cutoff: object = date.today()) -> ValidationReport:
    """Aggregate date anomalies without hiding residual chronology problems.

    Chronology warnings remain visible in ``warnings``/``issues``.  Future
    dates are hard errors because they cannot be imported as-is.  The function
    accepts either a collection of ``NormalizedTimeline`` values or a dataset
    exposing a ``timelines``/``transactions`` collection.
    """

    cutoff_date = coerce_excel_date(cutoff)
    timelines = tuple(timeline for row in _dataset_timelines(dataset) if (timeline := _timeline_from_row(row)) is not None)
    errors: list[str] = []
    warnings: list[str] = []
    issue_names: list[str] = []
    future_payment_count = 0
    payment_before_order_count = 0
    transactions_with_date_warnings = 0

    for index, timeline in enumerate(timelines, start=1):
        future_order = timeline.order > cutoff_date
        future_payments = sum(payment > cutoff_date for payment in timeline.payments)
        before_order = sum(payment < timeline.order for payment in timeline.payments)
        future_payment_count += future_payments
        payment_before_order_count += before_order
        # Sequence inversions can be legitimate late entry.  The residual
        # review count only covers dates that are impossible chronologically.
        if future_order or future_payments or before_order:
            transactions_with_date_warnings += 1
        if future_order:
            errors.append(f"timeline {index}: future_order")
        if future_payments:
            errors.append(f"timeline {index}: future_payment ({future_payments})")
        for issue in timeline.issues:
            if issue not in issue_names:
                issue_names.append(issue)
            if issue != "future_date":
                warnings.append(f"timeline {index}: {issue}")

    summary = (
        f"{'VALID' if not errors else 'INVALID'}: "
        f"{len(timelines)} timeline(s), "
        f"{future_payment_count} future payment(s), "
        f"{payment_before_order_count} payment(s) before order"
    )
    return ValidationReport(
        ok=not errors,
        errors=tuple(errors),
        warnings=tuple(warnings),
        issues=tuple(issue_names),
        future_payment_count=future_payment_count,
        payment_before_order_count=payment_before_order_count,
        transactions_with_date_warnings=transactions_with_date_warnings,
        summary=summary,
    )


@dataclass(frozen=True)
class MigrationPayment:
    no_transaksi: int
    nama_customer: str
    item: str
    urutan: int
    tanggal: date
    jumlah: int
    raw_tanggal: date


@dataclass(frozen=True)
class MigrationTransaction:
    no: int
    tanggal: date
    nama_raw: str
    nama_customer: str
    jenis: str
    item_raw: str
    item: str
    harga_beli: int
    harga_jual: int
    total_bayar: int
    sisa: int
    status: str
    timeline: NormalizedTimeline


@dataclass(frozen=True)
class MigrationCustomer:
    nama: str
    jumlah_transaksi: int
    nama_lama: tuple[str, ...]


@dataclass(frozen=True)
class MigrationDataset:
    transactions: tuple[MigrationTransaction, ...]
    customers: tuple[MigrationCustomer, ...]
    payments: tuple[MigrationPayment, ...]
    opening_funds: Mapping[str, int]
    validation: ValidationReport


def _money(value: object, *, field: str, transaction_no: int) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        raise ValueError(f"transaction {transaction_no}: invalid {field}")
    if isinstance(value, str):
        value = value.strip().replace("Rp", "").replace(" ", "").replace(".", "").replace(",", "")
    try:
        amount = int(value)
    except (TypeError, ValueError, OverflowError) as exc:
        raise ValueError(f"transaction {transaction_no}: invalid {field}: {value!r}") from exc
    if amount < 0:
        raise ValueError(f"transaction {transaction_no}: negative {field}")
    return amount


def _text(value: object, *, field: str, transaction_no: int) -> str:
    text = " ".join(str(value or "").split())
    if not text:
        raise ValueError(f"transaction {transaction_no}: missing {field}")
    return text


def _legacy_transactions(path: Path) -> dict[int, dict[str, object]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required")
    workbook = load_workbook(path, data_only=True, read_only=False)
    sheet = workbook["TRANSAKSI"]
    result: dict[int, dict[str, object]] = {}
    for values in sheet.iter_rows(min_row=3, values_only=True):
        if values[0] is None:
            continue
        number = int(values[0])
        result[number] = {
            "nama_raw": values[3],
            "jenis": str(values[5]).strip().lower(),
            "item": values[7],
        }
    return result


def _with_financial_validation(dataset: MigrationDataset, cutoff: date) -> ValidationReport:
    report = validate_dataset(dataset, cutoff=cutoff)
    actual = {
        "customers": len(dataset.customers),
        "transactions": len(dataset.transactions),
        "payments": len(dataset.payments),
        "harga_beli": sum(row.harga_beli for row in dataset.transactions),
        "harga_jual": sum(row.harga_jual for row in dataset.transactions),
        "payments_total": sum(row.jumlah for row in dataset.payments),
        "sisa": sum(row.sisa for row in dataset.transactions),
    }
    errors = list(report.errors)
    for key, expected in EXPECTED_CONTROLS.items():
        if actual[key] != expected:
            errors.append(f"control {key}: expected {expected}, got {actual[key]}")
    if sum(dataset.opening_funds.values()) != actual["sisa"]:
        errors.append("opening funds do not equal outstanding receivables")
    summary = (
        f"{'VALID' if not errors else 'INVALID'} WITH {report.payment_before_order_count} DATE WARNINGS: "
        f"{actual['customers']} nasabah, {actual['transactions']} transaksi, "
        f"{actual['payments']} pembayaran, piutang {_format_rupiah(actual['sisa'])}"
    )
    return replace(report, ok=not errors, errors=tuple(errors), summary=summary)


def build_dataset(
    r2_path: str | Path,
    legacy_preview_path: str | Path,
    cutoff: object,
) -> MigrationDataset:
    """Build the migration dataset from the client-cleaned r2 workbook."""

    if load_workbook is None:
        raise RuntimeError("openpyxl is required")
    cutoff_date = coerce_excel_date(cutoff)
    legacy = _legacy_transactions(Path(legacy_preview_path))
    workbook = load_workbook(Path(r2_path), data_only=True, read_only=False)
    sheet = workbook["UPDATE"]
    transactions: list[MigrationTransaction] = []
    payments: list[MigrationPayment] = []
    names_in_order: list[str] = []
    old_names: dict[str, set[str]] = {}

    for values in sheet.iter_rows(min_row=5, values_only=True):
        if values[0] is None:
            continue
        number = int(values[0])
        if number not in legacy:
            raise ValueError(f"transaction {number}: no legacy mapping")
        expected_number = len(transactions) + 1
        if number != expected_number:
            raise ValueError(f"transaction sequence: expected {expected_number}, got {number}")

        name = _text(values[2], field="customer", transaction_no=number)
        item_raw = " ".join(str(values[3] or "").split())
        if not item_raw and number not in ITEM_FIXES:
            raise ValueError(f"transaction {number}: missing item")
        legacy_row = legacy[number]
        jenis = str(legacy_row["jenis"])
        if jenis not in ALLOWED_JENIS:
            raise ValueError(f"transaction {number}: invalid jenis {jenis!r}")
        item = ITEM_FIXES.get(number, _text(legacy_row["item"], field="legacy item", transaction_no=number))

        raw_payment_dates: list[object] = []
        payment_amounts: list[int] = []
        for pair_start in range(7, 31, 2):
            raw_date = values[pair_start]
            amount = _money(values[pair_start + 1], field="payment", transaction_no=number)
            if amount:
                if raw_date is None:
                    raise DateNormalizationError(f"transaction {number}: payment has no date")
                raw_payment_dates.append(raw_date)
                payment_amounts.append(amount)
            elif raw_date is not None:
                raise ValueError(f"transaction {number}: payment date has no nominal")

        timeline = normalize_timeline(values[1], raw_payment_dates, cutoff=cutoff_date)
        harga_beli = _money(values[5], field="harga_beli", transaction_no=number)
        harga_jual = _money(values[6], field="harga_jual", transaction_no=number)
        total_bayar = sum(payment_amounts)
        sisa = harga_jual - total_bayar
        if sisa < 0:
            raise ValueError(f"transaction {number}: payment exceeds selling price")

        transaction = MigrationTransaction(
            no=number,
            tanggal=timeline.order,
            nama_raw=_text(legacy_row["nama_raw"], field="legacy customer", transaction_no=number),
            nama_customer=name,
            jenis=jenis,
            item_raw=item_raw,
            item=item,
            harga_beli=harga_beli,
            harga_jual=harga_jual,
            total_bayar=total_bayar,
            sisa=sisa,
            status="LUNAS" if sisa == 0 else "BELUM LUNAS",
            timeline=timeline,
        )
        transactions.append(transaction)
        if name not in old_names:
            names_in_order.append(name)
            old_names[name] = set()
        old_names[name].add(transaction.nama_raw)
        for index, (payment_date, raw_date, amount) in enumerate(
            zip(timeline.payments, timeline.raw[1:], payment_amounts), start=1
        ):
            payments.append(
                MigrationPayment(
                    no_transaksi=number,
                    nama_customer=name,
                    item=item,
                    urutan=index,
                    tanggal=payment_date,
                    jumlah=amount,
                    raw_tanggal=raw_date,
                )
            )

    counts = {name: 0 for name in names_in_order}
    for transaction in transactions:
        counts[transaction.nama_customer] += 1
    customers = tuple(
        MigrationCustomer(name, counts[name], tuple(sorted(old_names[name])))
        for name in names_in_order
    )
    provisional = MigrationDataset(
        transactions=tuple(transactions),
        customers=customers,
        payments=tuple(payments),
        opening_funds=dict(OPENING_FUNDS),
        validation=ValidationReport(ok=False),
    )
    return replace(provisional, validation=_with_financial_validation(provisional, cutoff_date))


_HEADER_FILL = "1F6E5A"
_SUBTLE_FILL = "DDEFE9"
_WARNING_FILL = "FFF1CC"


def _prepare_sheet(sheet: object, title: str, headers: Sequence[str]) -> None:
    sheet.append([title])
    sheet.append(list(headers))
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{sheet.cell(2, len(headers)).coordinate}"
    sheet.cell(1, 1).font = Font(bold=True, size=14, color="FFFFFF")
    sheet.cell(1, 1).fill = PatternFill("solid", fgColor=_HEADER_FILL)
    for cell in sheet[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _finish_sheet(sheet: object, widths: Sequence[int], money_columns: Iterable[int] = ()) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for column in money_columns:
        for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=3):
            cell[0].number_format = '#,##0'
    sheet.row_dimensions[2].height = 32


def write_preview(dataset: MigrationDataset, output_path: str | Path) -> None:
    """Write a formula-free review workbook suitable for the reseed step."""

    if Workbook is None:
        raise RuntimeError("openpyxl is required")
    if not dataset.validation.ok:
        raise ValueError("refusing to write an invalid migration dataset")
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    transactions_sheet = workbook.active
    transactions_sheet.title = "TRANSAKSI"
    _prepare_sheet(
        transactions_sheet,
        "PREVIEW MIGRASI R2 - nama dari klien, jenis dari preview lama, tanggal dinormalisasi",
        (
            "NO", "TGL", "THN", "NAMA_RAW", "NAMA_CUSTOMER", "JENIS",
            "ITEM_RAW", "ITEM_CLEAN", "H_BELI", "H_JUAL", "TOTAL_BAYAR",
            "SISA", "STATUS",
        ),
    )
    for row in dataset.transactions:
        transactions_sheet.append(
            (
                row.no, row.tanggal, row.tanggal.year, row.nama_raw,
                row.nama_customer, row.jenis, row.item_raw, row.item,
                row.harga_beli, row.harga_jual, row.total_bayar, row.sisa,
                row.status,
            )
        )
        transactions_sheet.cell(transactions_sheet.max_row, 2).number_format = "DD/MM/YYYY"
    _finish_sheet(
        transactions_sheet,
        (8, 13, 8, 22, 25, 16, 24, 24, 16, 16, 16, 16, 16),
        money_columns=(9, 10, 11, 12),
    )

    customer_sheet = workbook.create_sheet("CUSTOMER")
    _prepare_sheet(
        customer_sheet,
        "CUSTOMER R2 - identitas final mengikuti perapihan klien",
        ("NAMA_FINAL", "JML_TRX", "NAMA_LAMA"),
    )
    for row in dataset.customers:
        customer_sheet.append((row.nama, row.jumlah_transaksi, ", ".join(row.nama_lama)))
    _finish_sheet(customer_sheet, (28, 12, 52))

    payment_sheet = workbook.create_sheet("CICILAN")
    _prepare_sheet(
        payment_sheet,
        "DETAIL CICILAN R2 - tanggal final hasil normalisasi",
        ("NO_TRX", "NAMA_CUSTOMER", "ITEM", "URUTAN", "TGL_BAYAR", "NOMINAL"),
    )
    for row in dataset.payments:
        payment_sheet.append(
            (row.no_transaksi, row.nama_customer, row.item, row.urutan, row.tanggal, row.jumlah)
        )
        payment_sheet.cell(payment_sheet.max_row, 5).number_format = "DD/MM/YYYY"
    _finish_sheet(payment_sheet, (10, 26, 26, 10, 14, 16), money_columns=(6,))

    funds_sheet = workbook.create_sheet("SALDO_SUMBER_DANA")
    _prepare_sheet(
        funds_sheet,
        "SALDO AWAL SUMBER DANA - total harus sama dengan piutang berjalan",
        ("SUMBER_DANA", "SALDO_PIUTANG"),
    )
    for name, amount in dataset.opening_funds.items():
        funds_sheet.append((name, amount))
    _finish_sheet(funds_sheet, (24, 20), money_columns=(2,))

    validation_sheet = workbook.create_sheet("VALIDASI")
    _prepare_sheet(
        validation_sheet,
        dataset.validation.summary,
        ("NO_TRX", "STATUS", "FIELD", "URUTAN", "NILAI_ASLI", "NILAI_FINAL", "KETERANGAN"),
    )
    payments_by_transaction: dict[int, list[MigrationPayment]] = {}
    for payment in dataset.payments:
        payments_by_transaction.setdefault(payment.no_transaksi, []).append(payment)
    for transaction in dataset.transactions:
        for correction in transaction.timeline.corrections:
            index = int(correction["index"])
            validation_sheet.append(
                (
                    transaction.no,
                    "DIKOREKSI",
                    correction["field"],
                    None if index == 0 else index,
                    correction["raw"],
                    correction["normalized"],
                    correction["reason"],
                )
            )
        for payment in payments_by_transaction.get(transaction.no, []):
            if payment.tanggal < transaction.tanggal:
                validation_sheet.append(
                    (
                        transaction.no,
                        "PERLU CEK",
                        "payment",
                        payment.urutan,
                        payment.raw_tanggal,
                        payment.tanggal,
                        "payment_before_order",
                    )
                )
                for cell in validation_sheet[validation_sheet.max_row]:
                    cell.fill = PatternFill("solid", fgColor=_WARNING_FILL)
    for row in validation_sheet.iter_rows(min_row=3, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = "DD/MM/YYYY"
    _finish_sheet(validation_sheet, (10, 14, 14, 10, 16, 16, 28))

    workbook.properties.creator = "S&I Finance Solution"
    workbook.properties.created = datetime(2026, 8, 24)
    workbook.properties.modified = datetime(2026, 8, 24)
    workbook.save(output)


def verify_preview(path: str | Path) -> ValidationReport:
    """Reload a generated preview and independently verify its controls."""

    if load_workbook is None:
        raise RuntimeError("openpyxl is required")
    workbook = load_workbook(Path(path), data_only=True, read_only=False)
    expected_sheets = ["TRANSAKSI", "CUSTOMER", "CICILAN", "SALDO_SUMBER_DANA", "VALIDASI"]
    errors: list[str] = []
    if workbook.sheetnames != expected_sheets:
        errors.append(f"unexpected sheets: {workbook.sheetnames!r}")
        return ValidationReport(ok=False, errors=tuple(errors), summary="INVALID preview")

    transaction_rows = [row for row in workbook["TRANSAKSI"].iter_rows(min_row=3, values_only=True) if row[0] is not None]
    customer_rows = [row for row in workbook["CUSTOMER"].iter_rows(min_row=3, values_only=True) if row[0] is not None]
    payment_rows = [row for row in workbook["CICILAN"].iter_rows(min_row=3, values_only=True) if row[0] is not None]
    fund_rows = [row for row in workbook["SALDO_SUMBER_DANA"].iter_rows(min_row=3, values_only=True) if row[0] is not None]
    warning_rows = [
        row for row in workbook["VALIDASI"].iter_rows(min_row=3, values_only=True)
        if row[1] == "PERLU CEK"
    ]
    actual = {
        "customers": len(customer_rows),
        "transactions": len(transaction_rows),
        "payments": len(payment_rows),
        "harga_beli": sum(_money(row[8], field="harga_beli", transaction_no=int(row[0])) for row in transaction_rows),
        "harga_jual": sum(_money(row[9], field="harga_jual", transaction_no=int(row[0])) for row in transaction_rows),
        "payments_total": sum(_money(row[5], field="payment", transaction_no=int(row[0])) for row in payment_rows),
        "sisa": sum(_money(row[11], field="sisa", transaction_no=int(row[0])) for row in transaction_rows),
    }
    orders: dict[int, date] = {}
    sales: dict[int, int] = {}
    declared_payments: dict[int, int] = {}
    for row in transaction_rows:
        number = int(row[0])
        if number in orders:
            errors.append(f"duplicate transaction number: {number}")
        try:
            orders[number] = coerce_excel_date(row[1])
        except DateNormalizationError as error:
            errors.append(f"transaction {number}: {error}")
            continue
        if str(row[5]).strip().lower() not in ALLOWED_JENIS:
            errors.append(f"transaction {number}: invalid jenis {row[5]!r}")
        sales[number] = _money(row[9], field="harga_jual", transaction_no=number)
        declared_payments[number] = _money(row[10], field="total_bayar", transaction_no=number)

    payments_by_transaction: dict[int, int] = {}
    calculated_before_order = 0
    future_payment_count = 0
    warned_numbers: set[int] = set()
    today = date.today()
    for row in payment_rows:
        number = int(row[0])
        amount = _money(row[5], field="payment", transaction_no=number)
        payments_by_transaction[number] = payments_by_transaction.get(number, 0) + amount
        try:
            payment_date = coerce_excel_date(row[4])
        except DateNormalizationError as error:
            errors.append(f"transaction {number}: {error}")
            continue
        if payment_date > today:
            future_payment_count += 1
        if number not in orders:
            errors.append(f"payment references unknown transaction: {number}")
        elif payment_date < orders[number]:
            calculated_before_order += 1
            warned_numbers.add(number)
    for number, selling_price in sales.items():
        calculated_payment = payments_by_transaction.get(number, 0)
        if calculated_payment != declared_payments[number]:
            errors.append(
                f"transaction {number}: total payment expected {declared_payments[number]}, got {calculated_payment}"
            )
        row = transaction_rows[number - 1]
        declared_sisa = _money(row[11], field="sisa", transaction_no=number)
        if selling_price - calculated_payment != declared_sisa:
            errors.append(f"transaction {number}: inconsistent outstanding balance")
    if future_payment_count:
        errors.append(f"future payments: {future_payment_count}")
    for key, expected in EXPECTED_CONTROLS.items():
        if actual[key] != expected:
            errors.append(f"control {key}: expected {expected}, got {actual[key]}")
    if sum(_money(row[1], field="fund", transaction_no=0) for row in fund_rows) != actual["sisa"]:
        errors.append("fund balances do not equal outstanding receivables")
    if len(warning_rows) != 19:
        errors.append(f"expected 19 date warnings, got {len(warning_rows)}")
    warned_transactions = len({int(row[0]) for row in warning_rows})
    if warned_transactions != 17:
        errors.append(f"expected 17 transactions with date warnings, got {warned_transactions}")
    if calculated_before_order != len(warning_rows):
        errors.append(
            f"validation sheet has {len(warning_rows)} warnings, chronology has {calculated_before_order}"
        )
    if warned_numbers != {int(row[0]) for row in warning_rows}:
        errors.append("validation sheet transaction warnings do not match chronology")
    summary = (
        f"{'VALID' if not errors else 'INVALID'} WITH {len(warning_rows)} DATE WARNINGS: "
        f"{actual['customers']} nasabah, {actual['transactions']} transaksi, "
        f"{actual['payments']} pembayaran, piutang {_format_rupiah(actual['sisa'])}"
    )
    return ValidationReport(
        ok=not errors,
        errors=tuple(errors),
        warnings=tuple(f"transaction {row[0]}: payment_before_order" for row in warning_rows),
        issues=("payment_before_order",) if warning_rows else (),
        future_payment_count=future_payment_count,
        payment_before_order_count=len(warning_rows),
        transactions_with_date_warnings=warned_transactions,
        summary=summary,
    )


__all__ = [
    "ALLOWED_JENIS",
    "DateNormalizationError",
    "ITEM_FIXES",
    "MigrationCustomer",
    "MigrationDataset",
    "MigrationPayment",
    "MigrationTransaction",
    "NormalizedTimeline",
    "OPENING_FUNDS",
    "TimelineCandidate",
    "ValidationReport",
    "build_dataset",
    "coerce_excel_date",
    "date_candidates",
    "normalize_timeline",
    "validate_dataset",
    "verify_preview",
    "write_preview",
]
