#!/usr/bin/env python3
"""
Migrasi data dari PREVIEW_MIGRASI.xlsx ke Supabase.
Jalankan: uv run --with openpyxl,requests scripts/migrate_data.py
"""

import os, re, sys, uuid
from datetime import datetime, timezone
import openpyxl
import requests

# ── Kredensial ────────────────────────────────────────────────────────────────
SUPABASE_URL      = os.environ.get('SUPABASE_URL', '').rstrip('/')
SUPABASE_SVCKEY   = os.environ.get('SUPABASE_SERVICE_ROLE', '')

if not SUPABASE_URL or not SUPABASE_SVCKEY:
    # Coba baca dari .env
    env_path = os.path.join(os.path.dirname(__file__), '..', '.env')
    if os.path.exists(env_path):
        for line in open(env_path):
            line = line.strip()
            if line.startswith('SUPABASE_URL='):
                SUPABASE_URL = line.split('=', 1)[1].strip().rstrip('/')
            elif line.startswith('SUPABASE_SERVICE_ROLE='):
                SUPABASE_SVCKEY = line.split('=', 1)[1].strip()

if not SUPABASE_URL or not SUPABASE_SVCKEY:
    sys.exit('ERROR: SUPABASE_URL / SUPABASE_SERVICE_ROLE tidak ditemukan.')

HEADERS = {
    'apikey': SUPABASE_SVCKEY,
    'Authorization': f'Bearer {SUPABASE_SVCKEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=minimal',
}

def rest(method, table, data=None, params=None):
    url = f'{SUPABASE_URL}/rest/v1/{table}'
    r = getattr(requests, method)(url, headers=HEADERS, json=data, params=params)
    if r.status_code >= 300:
        print(f'  ERROR {r.status_code}: {r.text[:300]}')
        return None
    return r

def now_iso():
    return datetime.now(timezone.utc).isoformat()

# ── Baca Excel ────────────────────────────────────────────────────────────────
XLSX = os.path.join(os.path.dirname(__file__), '..', 'ref', 'PREVIEW_MIGRASI.xlsx')
wb = openpyxl.load_workbook(XLSX)

def parse_tgl(val, fallback_year=2020):
    if isinstance(val, datetime):
        return val
    if not val:
        return datetime(fallback_year, 1, 1)
    s = str(val).strip()
    for fmt in ['%d/%m/%Y', '%d/%m/%y', '%d-%m-%Y']:
        try:
            return datetime.strptime(s, fmt)
        except:
            pass
    try:
        parts = re.split(r'[/\-]', s)
        if len(parts) == 3:
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            if y < 100: y += 2000
            return datetime(y, m, d)
    except:
        pass
    return datetime(fallback_year, 1, 1)

# Sheet TRANSAKSI
ws_t = wb['TRANSAKSI']
transactions = []
for row in ws_t.iter_rows(min_row=3, values_only=True):
    if row[0] is None:
        break
    thn = row[2] or 2020
    tgl_dt = parse_tgl(row[1], int(thn))
    transactions.append({
        'no_urut':       row[0],
        'nama_customer': str(row[4]).strip() if row[4] else str(row[3]).strip(),
        'jenis':         str(row[5]).strip() if row[5] else 'barang',
        'nama_barang':   str(row[7]).strip() if row[7] else (str(row[6]).strip() if row[6] else ''),
        'h_beli':        int(row[8]) if row[8] else 0,
        'h_jual':        int(row[9]) if row[9] else 0,
        'tanggal':       tgl_dt,
    })

# Sheet CICILAN
ws_c = wb['CICILAN']
cicilan_map = {}  # no_urut -> list[{tgl, nominal}]
for row in ws_c.iter_rows(min_row=3, values_only=True):
    if row[0] is None:
        break
    no_trx = row[0]
    nominal = row[5]
    if not nominal or nominal == 0:
        continue
    tgl_dt = parse_tgl(row[4])
    cicilan_map.setdefault(no_trx, []).append({
        'tgl': tgl_dt,
        'nominal': int(nominal),
    })

print(f'Transaksi: {len(transactions)}, Cicilan: {sum(len(v) for v in cicilan_map.values())}')

# ── Ambil admin profile id ─────────────────────────────────────────────────────
r = requests.get(
    f'{SUPABASE_URL}/rest/v1/profiles',
    headers=HEADERS,
    params={'role': 'eq.admin', 'limit': '1'},
)
profiles = r.json() if r.status_code == 200 else []
admin_id = profiles[0]['id'] if profiles else None
print(f'Admin profile id: {admin_id}')

# ── Cek apakah sudah ada data (idempotent guard) ───────────────────────────────
r = requests.get(
    f'{SUPABASE_URL}/rest/v1/customers',
    headers=HEADERS,
    params={'limit': '1'},
)
existing = r.json() if r.status_code == 200 else []
if existing:
    print(f'PERINGATAN: tabel customers sudah ada {len(existing)}+ data.')
    ans = input('Lanjut insert? Data duplikat akan di-skip karena pakai upsert. (y/N): ').strip().lower()
    if ans != 'y':
        sys.exit('Dibatalkan.')

# ── Build customer list ────────────────────────────────────────────────────────
print('\n--- Membuat customers ---')
# Kumpulkan nama unik, urutkan sesuai kemunculan pertama
seen = {}
for t in transactions:
    nama = t['nama_customer']
    if nama not in seen:
        seen[nama] = t['tanggal']

customer_rows = []
customer_id_map = {}  # nama -> uuid
for nama, first_tgl in seen.items():
    cid = str(uuid.uuid4())
    customer_id_map[nama] = cid
    ts = first_tgl.replace(tzinfo=timezone.utc).isoformat()
    customer_rows.append({
        'id':         cid,
        'nama':       nama,
        'created_by': admin_id,
        'created_at': ts,
        'updated_at': ts,
    })

# Upsert customers dalam batch
BATCH = 50
for i in range(0, len(customer_rows), BATCH):
    chunk = customer_rows[i:i+BATCH]
    r = requests.post(
        f'{SUPABASE_URL}/rest/v1/customers',
        headers={**HEADERS, 'Prefer': 'resolution=merge-duplicates,return=minimal'},
        json=chunk,
    )
    status = 'OK' if r.status_code < 300 else f'ERROR {r.status_code}: {r.text[:200]}'
    print(f'  Customers [{i+1}-{i+len(chunk)}]: {status}')

print(f'Total customers: {len(customer_rows)}')

# ── Build purchases + payments ────────────────────────────────────────────────
print('\n--- Membuat purchases & payments ---')
purchase_rows = []
payment_rows  = []

for t in transactions:
    pid = str(uuid.uuid4())
    cid = customer_id_map.get(t['nama_customer'])
    if not cid:
        print(f'  SKIP: customer tidak ditemukan: {t["nama_customer"]}')
        continue

    ts_beli = t['tanggal'].replace(tzinfo=timezone.utc).isoformat()
    purchase_rows.append({
        'id':          pid,
        'customer_id': cid,
        'nama_barang': t['nama_barang'],
        'jenis':       t['jenis'],
        'harga_jual':  t['h_jual'],
        'harga_beli':  t['h_beli'] if t['h_beli'] > 0 else None,
        'tanggal_beli': t['tanggal'].strftime('%Y-%m-%d'),
        'created_by':  admin_id,
        'created_at':  ts_beli,
        'updated_at':  ts_beli,
    })

    # Cicilan → payments
    for c in cicilan_map.get(t['no_urut'], []):
        ts_bayar = c['tgl'].replace(tzinfo=timezone.utc).isoformat()
        payment_rows.append({
            'id':                 str(uuid.uuid4()),
            'customer_id':        cid,
            'jumlah':             c['nominal'],
            'tanggal_bayar':      c['tgl'].strftime('%Y-%m-%d'),
            'metode':             'tunai',
            'sumber':             'admin',
            'status_verifikasi':  'verified',
            'created_by':         admin_id,
            'created_at':         ts_bayar,
            'updated_at':         ts_bayar,
        })

# Upsert purchases
for i in range(0, len(purchase_rows), BATCH):
    chunk = purchase_rows[i:i+BATCH]
    r = requests.post(
        f'{SUPABASE_URL}/rest/v1/purchases',
        headers={**HEADERS, 'Prefer': 'resolution=merge-duplicates,return=minimal'},
        json=chunk,
    )
    status = 'OK' if r.status_code < 300 else f'ERROR {r.status_code}: {r.text[:200]}'
    print(f'  Purchases [{i+1}-{i+len(chunk)}]: {status}')

print(f'Total purchases: {len(purchase_rows)}')

# Upsert payments
for i in range(0, len(payment_rows), BATCH):
    chunk = payment_rows[i:i+BATCH]
    r = requests.post(
        f'{SUPABASE_URL}/rest/v1/payments',
        headers={**HEADERS, 'Prefer': 'resolution=merge-duplicates,return=minimal'},
        json=chunk,
    )
    status = 'OK' if r.status_code < 300 else f'ERROR {r.status_code}: {r.text[:200]}'
    print(f'  Payments [{i+1}-{i+len(chunk)}]: {status}')

print(f'Total payments: {len(payment_rows)}')
print('\nMigrasi selesai.')
