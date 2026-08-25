#!/usr/bin/env python3
"""Build, inspect, back up, and atomically apply the r2 reseed payload."""

from __future__ import annotations

import argparse
from datetime import date, datetime, timezone
import json
import os
from pathlib import Path
import sys
from typing import Any, Mapping
import uuid

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - supplied by the migration command
    load_workbook = None

try:
    import requests
except ImportError:  # pragma: no cover - supplied by the migration command
    requests = None

try:
    from scripts.migration_r2 import verify_preview
except ImportError:  # Direct execution places scripts/ on sys.path.
    from migration_r2 import verify_preview


ROOT = Path(__file__).resolve().parents[1]
CONFIRMATION_PHRASE = "RESET-BUSINESS-DATA"
BUSINESS_TABLES = (
    "fund_ledger_entries",
    "budget_entries",
    "payments",
    "purchases",
    "customers",
    "fund_sources",
)
SOURCE_IDS = {
    "Sandi": "00000000-0000-4000-8000-000000000001",
    "Ika": "00000000-0000-4000-8000-000000000002",
}
SOURCE_COLORS = {"Sandi": "green", "Ika": "gold"}
RESEED_NAMESPACE = uuid.UUID("f7896812-1702-4c12-92a0-829b72302663")
OPENING_DATE = date(2026, 8, 24)


class SafetyError(RuntimeError):
    """Raised when a reseed safety condition or reconciliation fails."""


def _stable_id(key: str) -> str:
    return str(uuid.uuid5(RESEED_NAMESPACE, key))


def _date_string(value: object, *, field: str) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        try:
            return date.fromisoformat(value).isoformat()
        except ValueError as error:
            raise SafetyError(f"Tanggal {field} tidak valid: {value!r}") from error
    raise SafetyError(f"Tanggal {field} tidak valid: {value!r}")


def _timestamp(value: object, *, field: str) -> str:
    return f"{_date_string(value, field=field)}T00:00:00+00:00"


def _money(value: object, *, field: str) -> int:
    if value is None or isinstance(value, bool):
        raise SafetyError(f"Nominal {field} tidak valid: {value!r}")
    try:
        amount = int(value)
    except (TypeError, ValueError, OverflowError) as error:
        raise SafetyError(f"Nominal {field} tidak valid: {value!r}") from error
    if amount < 0:
        raise SafetyError(f"Nominal {field} tidak boleh negatif")
    return amount


def _rows(sheet: Any) -> list[tuple[object, ...]]:
    return [
        row
        for row in sheet.iter_rows(min_row=3, values_only=True)
        if row[0] is not None
    ]


def build_reseed_payload(preview_path: str | Path) -> dict[str, Any]:
    """Convert a verified preview workbook into a deterministic RPC payload."""

    if load_workbook is None:
        raise RuntimeError("openpyxl diperlukan untuk membaca preview")
    preview = Path(preview_path)
    if not preview.is_file():
        raise SafetyError(f"Preview tidak ditemukan: {preview}")
    verification = verify_preview(preview)
    if not verification.ok:
        raise SafetyError("Preview gagal diverifikasi: " + "; ".join(verification.errors))
    workbook = load_workbook(preview, data_only=True, read_only=True)
    expected_sheets = {
        "TRANSAKSI",
        "CUSTOMER",
        "CICILAN",
        "SALDO_SUMBER_DANA",
        "VALIDASI",
    }
    if set(workbook.sheetnames) != expected_sheets:
        raise SafetyError(f"Sheet preview tidak sesuai: {workbook.sheetnames!r}")

    transaction_rows = _rows(workbook["TRANSAKSI"])
    customer_rows = _rows(workbook["CUSTOMER"])
    payment_rows = _rows(workbook["CICILAN"])
    fund_rows = _rows(workbook["SALDO_SUMBER_DANA"])

    customer_ids = {
        str(row[0]).strip(): _stable_id(f"customer:{str(row[0]).strip()}")
        for row in customer_rows
    }
    earliest_dates: dict[str, str] = {}
    purchase_ids: dict[int, str] = {}
    purchases: list[dict[str, Any]] = []
    for row in transaction_rows:
        number = int(row[0])
        purchased_on = _date_string(row[1], field=f"transaksi {number}")
        customer_name = str(row[4]).strip()
        if customer_name not in customer_ids:
            raise SafetyError(f"Transaksi {number} merujuk nasabah tak dikenal")
        earliest_dates[customer_name] = min(
            purchased_on,
            earliest_dates.get(customer_name, purchased_on),
        )
        purchase_id = _stable_id(f"purchase:{number}")
        purchase_ids[number] = purchase_id
        purchases.append(
            {
                "id": purchase_id,
                "customer_id": customer_ids[customer_name],
                "nama_barang": str(row[7]).strip(),
                "jenis": str(row[5]).strip().lower(),
                "harga_jual": _money(row[9], field=f"harga jual transaksi {number}"),
                "harga_beli": _money(row[8], field=f"harga beli transaksi {number}"),
                "tanggal_beli": purchased_on,
                "catatan": f"Migrasi r2 - transaksi {number}",
                "fund_source_id": None,
                "created_by": None,
                "created_at": _timestamp(row[1], field=f"transaksi {number}"),
                "updated_at": _timestamp(row[1], field=f"transaksi {number}"),
                "deleted_at": None,
            }
        )

    customers = []
    for row in customer_rows:
        name = str(row[0]).strip()
        created_on = earliest_dates.get(name, OPENING_DATE.isoformat())
        customers.append(
            {
                "id": customer_ids[name],
                "nama": name,
                "no_hp": None,
                "alamat": None,
                "catatan": None,
                "is_archived": False,
                "auth_user_id": None,
                "created_by": None,
                "created_at": f"{created_on}T00:00:00+00:00",
                "updated_at": f"{created_on}T00:00:00+00:00",
                "deleted_at": None,
            }
        )

    payments: list[dict[str, Any]] = []
    for row in payment_rows:
        number = int(row[0])
        sequence = int(row[3])
        customer_name = str(row[1]).strip()
        if number not in purchase_ids or customer_name not in customer_ids:
            raise SafetyError(f"Cicilan transaksi {number} memiliki referensi tak dikenal")
        payment_date = _date_string(row[4], field=f"cicilan transaksi {number}")
        payments.append(
            {
                "id": _stable_id(f"payment:{number}:{sequence}"),
                "customer_id": customer_ids[customer_name],
                "jumlah": _money(row[5], field=f"cicilan transaksi {number}"),
                "tanggal_bayar": payment_date,
                "metode": "tunai",
                "catatan": f"Migrasi r2 - transaksi {number}, cicilan {sequence}",
                "sumber": "admin",
                "status_verifikasi": "verified",
                "bukti_foto_url": None,
                "fund_source_id": None,
                "created_by": None,
                "created_at": f"{payment_date}T00:00:00+00:00",
                "updated_at": f"{payment_date}T00:00:00+00:00",
                "deleted_at": None,
            }
        )

    opening_funds = {
        str(row[0]).strip(): _money(row[1], field=f"saldo {row[0]}")
        for row in fund_rows
    }
    if set(opening_funds) != set(SOURCE_IDS):
        raise SafetyError(f"Sumber dana harus tepat Sandi dan Ika: {sorted(opening_funds)}")
    fund_sources = [
        {
            "id": SOURCE_IDS[name],
            "nama": name,
            "color_key": SOURCE_COLORS[name],
            "is_active": True,
            "created_by": None,
            "created_at": f"{OPENING_DATE.isoformat()}T00:00:00+00:00",
            "updated_at": f"{OPENING_DATE.isoformat()}T00:00:00+00:00",
            "deleted_at": None,
        }
        for name in ("Sandi", "Ika")
    ]
    ledger = [
        {
            "id": _stable_id(f"opening-fund:{name}"),
            "fund_source_id": SOURCE_IDS[name],
            "tanggal": OPENING_DATE.isoformat(),
            "tipe": "saldo_awal",
            "jumlah_delta": opening_funds[name],
            "reference_type": "migration",
            "reference_id": None,
            "transfer_group_id": None,
            "catatan": "Saldo awal migrasi r2",
            "created_by": None,
            "created_at": f"{OPENING_DATE.isoformat()}T00:00:00+00:00",
            "updated_at": f"{OPENING_DATE.isoformat()}T00:00:00+00:00",
            "deleted_at": None,
        }
        for name in ("Sandi", "Ika")
    ]

    controls = {
        "customers": len(customers),
        "purchases": len(purchases),
        "payments": len(payments),
        "fund_sources": len(fund_sources),
        "fund_ledger_entries": len(ledger),
        "harga_beli": sum(row["harga_beli"] for row in purchases),
        "harga_jual": sum(row["harga_jual"] for row in purchases),
        "payments_total": sum(row["jumlah"] for row in payments),
        "outstanding": sum(row["harga_jual"] for row in purchases)
        - sum(row["jumlah"] for row in payments),
        "fund_balance_total": sum(row["jumlah_delta"] for row in ledger),
    }
    required = {
        "customers": 46,
        "purchases": 249,
        "payments": 1069,
        "fund_sources": 2,
        "fund_ledger_entries": 2,
        "harga_beli": 877_769_000,
        "harga_jual": 971_417_500,
        "payments_total": 854_692_500,
        "outstanding": 116_725_000,
        "fund_balance_total": 116_725_000,
    }
    if controls != required:
        raise SafetyError(f"Kontrol preview tidak sesuai: {controls!r}")
    return {
        "controls": controls,
        "fund_sources": fund_sources,
        "customers": customers,
        "purchases": purchases,
        "payments": payments,
        "fund_ledger_entries": ledger,
        "budget_entries": [],
    }


class SupabaseRestClient:
    """Small service-role REST client with explicit read and RPC surfaces."""

    def __init__(self, url: str, service_key: str, page_size: int = 1000):
        if requests is None:
            raise RuntimeError("requests diperlukan untuk akses Supabase")
        self.base_url = url.rstrip("/")
        self.page_size = page_size
        self.headers = {
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Content-Type": "application/json",
        }

    def fetch_all(self, table: str) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        offset = 0
        while True:
            response = requests.get(
                f"{self.base_url}/rest/v1/{table}",
                headers=self.headers,
                params={"select": "*", "order": "id.asc", "limit": self.page_size, "offset": offset},
                timeout=60,
            )
            response.raise_for_status()
            page = response.json()
            if not isinstance(page, list):
                raise SafetyError(f"Respons {table} bukan daftar")
            rows.extend(page)
            if len(page) < self.page_size:
                return rows
            offset += self.page_size

    def rpc(self, name: str, payload: dict[str, Any]) -> dict[str, Any]:
        response = requests.post(
            f"{self.base_url}/rest/v1/rpc/{name}",
            headers=self.headers,
            json={"p_payload": payload},
            timeout=120,
        )
        response.raise_for_status()
        result = response.json()
        if not isinstance(result, dict):
            raise SafetyError(f"Respons RPC {name} tidak valid")
        return result


def backup_tables(
    client: Any,
    backup_root: str | Path = ROOT / "backups",
    *,
    timestamp: str | None = None,
) -> Path:
    """Export every business row, including soft-deleted rows, before apply."""

    stamp = timestamp or datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    destination = Path(backup_root) / f"reseed-{stamp}"
    destination.mkdir(parents=True, exist_ok=False)
    for table in BUSINESS_TABLES:
        rows = client.fetch_all(table)
        (destination / f"{table}.json").write_text(
            json.dumps(rows, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
    return destination


def _active(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row for row in rows if row.get("deleted_at") is None]


def _remote_metrics(tables: Mapping[str, list[dict[str, Any]]]) -> dict[str, Any]:
    customers = _active(tables["customers"])
    purchases = _active(tables["purchases"])
    payments = [
        row
        for row in _active(tables["payments"])
        if row.get("status_verifikasi") == "verified"
    ]
    sources = _active(tables["fund_sources"])
    ledger = _active(tables["fund_ledger_entries"])
    customer_ids = {row["id"] for row in customers}
    source_ids = {row["id"] for row in sources}
    orphans = sum(row.get("customer_id") not in customer_ids for row in purchases)
    orphans += sum(row.get("customer_id") not in customer_ids for row in payments)
    orphans += sum(row.get("fund_source_id") not in source_ids for row in ledger)
    orphans += sum(
        row.get("fund_source_id") is not None and row.get("fund_source_id") not in source_ids
        for row in purchases + payments
    )
    balances = {
        source["nama"]: sum(
            int(row["jumlah_delta"])
            for row in ledger
            if row["fund_source_id"] == source["id"]
        )
        + sum(
            int(row["harga_jual"])
            for row in purchases
            if row.get("fund_source_id") == source["id"]
        )
        - sum(
            int(row["jumlah"])
            for row in payments
            if row.get("fund_source_id") == source["id"]
        )
        for source in sources
    }
    outstanding = sum(int(row["harga_jual"]) for row in purchases) - sum(
        int(row["jumlah"]) for row in payments
    )
    return {
        "customers": len(customers),
        "purchases": len(purchases),
        "payments": len(payments),
        "fund_sources": len(sources),
        "fund_ledger_entries": len(ledger),
        "budget_entries": len(_active(tables["budget_entries"])),
        "harga_beli": sum(int(row.get("harga_beli") or 0) for row in purchases),
        "harga_jual": sum(int(row["harga_jual"]) for row in purchases),
        "payments_total": sum(int(row["jumlah"]) for row in payments),
        "outstanding": outstanding,
        "fund_balances": balances,
        "fund_balance_total": sum(balances.values()),
        "orphans": orphans,
    }


def reconcile_remote(
    client: Any,
    expected: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Read all business tables and optionally enforce the r2 target controls."""

    tables = {table: client.fetch_all(table) for table in BUSINESS_TABLES}
    metrics = _remote_metrics(tables)
    errors: list[str] = []
    if metrics["orphans"]:
        errors.append(f"orphan: {metrics['orphans']}")
    if expected is not None:
        controls = expected["controls"]
        for key in (
            "customers",
            "purchases",
            "payments",
            "fund_sources",
            "fund_ledger_entries",
            "harga_beli",
            "harga_jual",
            "payments_total",
            "outstanding",
            "fund_balance_total",
        ):
            if metrics[key] != controls[key]:
                errors.append(f"{key}: target {controls[key]}, remote {metrics[key]}")
        if metrics["budget_entries"] != 0:
            errors.append(f"budget_entries: target 0, remote {metrics['budget_entries']}")
        expected_balances = {
            source["nama"]: next(
                row["jumlah_delta"]
                for row in expected["fund_ledger_entries"]
                if row["fund_source_id"] == source["id"]
            )
            for source in expected["fund_sources"]
        }
        if metrics["fund_balances"] != expected_balances:
            errors.append(
                f"fund_balances: target {expected_balances}, remote {metrics['fund_balances']}"
            )
        if metrics["fund_balance_total"] != metrics["outstanding"]:
            errors.append("total saldo sumber dana tidak sama dengan piutang")
    metrics["ok"] = not errors
    metrics["errors"] = errors
    if errors:
        raise SafetyError("Rekonsiliasi gagal: " + "; ".join(errors))
    return metrics


def run_reseed(
    preview_path: str | Path,
    *,
    client: Any | None = None,
    apply: bool = False,
    confirmation: str | None = None,
    backup_root: str | Path = ROOT / "backups",
) -> dict[str, Any]:
    """Dry-run by default; on apply, require backup then one atomic RPC."""

    payload = build_reseed_payload(preview_path)
    if not apply:
        return {"mode": "dry-run", "controls": payload["controls"]}
    if confirmation != CONFIRMATION_PHRASE:
        raise SafetyError(f"Apply memerlukan --confirm-reset {CONFIRMATION_PHRASE}")
    if client is None:
        raise SafetyError("Klien Supabase diperlukan untuk apply")
    backup_path = backup_tables(client, backup_root)
    rpc_result = client.rpc("reseed_business_data_v2", payload)
    reconciliation = reconcile_remote(client, expected=payload)
    return {
        "mode": "apply",
        "backup": str(backup_path),
        "rpc": rpc_result,
        "reconciliation": reconciliation,
    }


def _load_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.is_file():
        return values
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def _remote_client() -> SupabaseRestClient:
    file_env = _load_env(ROOT / ".env")
    url = os.environ.get("SUPABASE_URL") or file_env.get("SUPABASE_URL")
    service_key = (
        os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
        or os.environ.get("SUPABASE_SERVICE_ROLE")
        or file_env.get("SUPABASE_SERVICE_ROLE_KEY")
        or file_env.get("SUPABASE_SERVICE_ROLE")
    )
    if not url or not service_key:
        raise SafetyError("SUPABASE_URL dan SUPABASE_SERVICE_ROLE_KEY diperlukan")
    return SupabaseRestClient(url, service_key)


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--preview",
        type=Path,
        default=ROOT / "ref" / "PREVIEW_MIGRASI_R2.xlsx",
    )
    parser.add_argument("--remote-check", action="store_true")
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--confirm-reset")
    parser.add_argument("--backup-root", type=Path, default=ROOT / "backups")
    return parser


def _print_metrics(title: str, metrics: Mapping[str, Any]) -> None:
    print(title)
    for key in (
        "customers",
        "purchases",
        "payments",
        "harga_beli",
        "harga_jual",
        "payments_total",
        "outstanding",
        "fund_balance_total",
    ):
        if key in metrics:
            print(f"  {key}: {metrics[key]:,}".replace(",", "."))


def main(argv: list[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    if args.apply and args.remote_check:
        print("ERROR: --apply dan --remote-check tidak dapat digabung", file=sys.stderr)
        return 2
    try:
        payload = build_reseed_payload(args.preview)
        _print_metrics("Target preview r2:", payload["controls"])
        if args.remote_check:
            current = reconcile_remote(_remote_client())
            _print_metrics("Data Supabase saat ini (read-only):", current)
            print("Tidak ada perubahan Supabase.")
            return 0
        result = run_reseed(
            args.preview,
            client=_remote_client() if args.apply else None,
            apply=args.apply,
            confirmation=args.confirm_reset,
            backup_root=args.backup_root,
        )
        if args.apply:
            print(f"Backup: {result['backup']}")
            _print_metrics("Rekonsiliasi setelah reseed:", result["reconciliation"])
            print("Reseed selesai dan terverifikasi.")
        else:
            print("Tidak ada perubahan Supabase.")
        return 0
    except (OSError, ValueError, SafetyError) as error:
        print(f"ERROR: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
